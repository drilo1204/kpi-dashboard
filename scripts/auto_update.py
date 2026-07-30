#!/usr/bin/env python3
# Auto-Update fuer KPI-Dashboard. Wird woechentlich von GitHub Actions aufgerufen.
# Flow: Drive laden -> Excel -> JSON -> Anthropic mit RULES.md -> neue data.js -> validieren.
# Push und Mail uebernimmt der Workflow.

import io
import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

import anthropic
import openpyxl
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_JS = REPO_ROOT / "data.js"
RULES_MD = REPO_ROOT / "RULES.md"
DRIVE_FOLDER_ID = os.environ.get("DRIVE_FOLDER_ID", "1x1RW-0GIZw8hZs4ZyxftKRcbFCbHhkzf")
# claude-sonnet-4-5 = stabil, ausreichend fuer diese Aufgabe.
# Kann per Repo-Variable ANTHROPIC_MODEL ueberschrieben werden.
ANTHROPIC_MODEL = os.environ.get("ANTHROPIC_MODEL", "claude-sonnet-4-5")
MAX_ROWS_PER_SHEET = 300
MAX_TOKENS_OUT = 16000


def log(msg: str) -> None:
    print(msg, flush=True)


def set_output(key: str, value: str) -> None:
    out = os.environ.get("GITHUB_OUTPUT")
    if not out:
        return
    with open(out, "a", encoding="utf-8") as fh:
        fh.write(f"{key}={value}\n")


def get_drive_service():
    creds_json = os.environ["GOOGLE_SERVICE_ACCOUNT_JSON"]
    info = json.loads(creds_json)
    creds = service_account.Credentials.from_service_account_info(
        info, scopes=["https://www.googleapis.com/auth/drive.readonly"]
    )
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def list_xlsx_files(drive, folder_id):
    mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    q = f"'{folder_id}' in parents and mimeType='{mime}' and trashed=false"
    result = (
        drive.files()
        .list(q=q, fields="files(id, name, modifiedTime)", pageSize=100)
        .execute()
    )
    return result.get("files", [])


def download_file(drive, file_id: str) -> io.BytesIO:
    request = drive.files().get_media(fileId=file_id)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    buf.seek(0)
    return buf


def excel_to_dict(buf: io.BytesIO, filename: str) -> dict:
    wb = openpyxl.load_workbook(buf, data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            clean = []
            for cell in row:
                if hasattr(cell, "isoformat"):
                    clean.append(cell.isoformat())
                elif cell is None:
                    clean.append(None)
                else:
                    clean.append(cell)
            rows.append(clean)
        sheets[name] = rows
    return {"filename": filename, "sheets": sheets}


def format_excel_section(excel_data: list) -> str:
    parts = []
    for f in excel_data:
        parts.append(f"\n### {f['filename']}")
        for sheet_name, rows in f["sheets"].items():
            parts.append(f"\n#### Sheet: {sheet_name}  ({len(rows)} Zeilen)")
            shown = rows[:MAX_ROWS_PER_SHEET]
            for row in shown:
                parts.append(repr(row))
            if len(rows) > MAX_ROWS_PER_SHEET:
                parts.append(f"... ({len(rows) - MAX_ROWS_PER_SHEET} weitere Zeilen abgeschnitten)")
    return "\n".join(parts)


def build_prompt(rules: str, current_data_js: str, excel_data: list) -> str:
    heute = datetime.now(timezone.utc).strftime("%d.%m.%Y")
    excel_section = format_excel_section(excel_data)
    header = (
        "Du aktualisierst die data.js des KPI-Dashboards von Brunner Mobil. "
        "Halte dich strikt an die RULES.md.\n\n"
        "WICHTIG - Ausgabeformat:\n"
        "- Gib NUR den vollstaendigen neuen data.js-Inhalt zurueck.\n"
        "- Kein Markdown-Fence, kein Vor- oder Nachtext, keine Erklaerung.\n"
        "- Behalte die vorhandene Struktur bei (alle KPI-Konstanten, Reihenfolge, Kommentare mit Rechenweg).\n"
        f"- Aktualisiere `letzteAktualisierung` in DASHBOARD_CONFIG auf das heutige Datum: {heute}.\n"
        "- Wenn ein Wert unklar oder die Datengrundlage duenn ist: bestehenden Wert behalten und in Kommentar begruenden.\n"
        "- Laufende Perioden (aktueller Monat/Woche) NIE als Trend-Datenpunkt aufnehmen (siehe RULES 0.2).\n"
    )
    footer = (
        "\nGib jetzt die neue data.js aus - beginnend mit dem Kommentar '// ===...' "
        "und endend mit dem letzten schliessenden '};'.\n"
    )
    return (
        header
        + "\n# RULES.md\n"
        + rules
        + "\n\n# Aktuelle data.js\n```\n"
        + current_data_js
        + "\n```\n\n# Excel-Rohdaten aus dem Drive-Ordner\n"
        + excel_section
        + footer
    )


def call_anthropic(prompt: str) -> tuple[str, str, dict]:
    """Ruft Anthropic auf und liefert (text, stop_reason, usage_dict).
    Bei API-Fehler wird die Exception nach oben durchgereicht."""
    client = anthropic.Anthropic(api_key=os.environ["ANTHROPIC_API_KEY"])
    response = client.messages.create(
        model=ANTHROPIC_MODEL,
        max_tokens=MAX_TOKENS_OUT,
        messages=[{"role": "user", "content": prompt}],
    )
    stop_reason = getattr(response, "stop_reason", "unknown")
    usage = {}
    if hasattr(response, "usage") and response.usage is not None:
        usage = {
            "input_tokens": getattr(response.usage, "input_tokens", None),
            "output_tokens": getattr(response.usage, "output_tokens", None),
        }

    block_types = [getattr(b, "type", "unknown") for b in response.content]
    log(f"      Response-Block-Typen: {block_types}")

    text = "".join(
        getattr(block, "text", "") for block in response.content
        if getattr(block, "type", None) == "text"
    )
    return text.strip(), stop_reason, usage


def strip_code_fence(text: str) -> str:
    t = text.strip()
    if t.startswith("```"):
        first_nl = t.find("\n")
        if first_nl != -1:
            t = t[first_nl + 1 :]
    if t.endswith("```"):
        t = t[: t.rfind("```")].rstrip()
    return t.strip()


def validate_data_js(text: str) -> tuple[bool, str]:
    if not text.startswith("//"):
        return False, "Startet nicht mit Kommentar"
    required = [
        "DASHBOARD_CONFIG",
        "KPI_GOOGLE",
        "KPI_REFERENZ",
        "KPI_NPS",
        "KPI_CONVERSION",
        "KPI_VERTRAGSEINGAENGE_WOCHE",
        "KPI_GRAFIK",
    ]
    for r in required:
        if r not in text:
            return False, f"'{r}' fehlt im Output"
    if text.count("{") != text.count("}"):
        return False, "Geschweifte Klammern unbalanciert"
    if text.count("[") != text.count("]"):
        return False, "Eckige Klammern unbalanciert"
    return True, "OK"


def main() -> int:
    log(f"[1/7] Google Drive verbinden (Ordner {DRIVE_FOLDER_ID}) ...")
    drive = get_drive_service()

    log("[2/7] Excel-Dateien im Ordner listen ...")
    files = list_xlsx_files(drive, DRIVE_FOLDER_ID)
    log(f"      {len(files)} .xlsx-Dateien gefunden.")
    if not files:
        log("      Kein Input. Nichts zu tun.")
        set_output("changed", "false")
        set_output("reason", "keine-excel-im-drive")
        return 0

    log("[3/7] Excel-Dateien herunterladen und parsen ...")
    excel_data = []
    for f in files:
        log(f"      - {f['name']}  (modified {f.get('modifiedTime', '?')})")
        buf = download_file(drive, f["id"])
        excel_data.append(excel_to_dict(buf, f["name"]))

    log("[4/7] RULES.md und aktuelle data.js lesen ...")
    if not RULES_MD.exists():
        log("      FEHLER: RULES.md nicht gefunden.")
        set_output("changed", "false")
        set_output("reason", "rules-md-fehlt")
        return 2
    rules = RULES_MD.read_text(encoding="utf-8")
    current_data_js = DATA_JS.read_text(encoding="utf-8")

    log(f"[5/7] Anthropic aufrufen (Modell {ANTHROPIC_MODEL}) ...")
    prompt = build_prompt(rules, current_data_js, excel_data)
    log(f"      Prompt-Groesse: {len(prompt)} Zeichen (grob {len(prompt)//4} Tokens).")
    try:
        raw_text, stop_reason, usage = call_anthropic(prompt)
    except Exception as e:
        log(f"      FEHLER beim Anthropic-Call: {type(e).__name__}: {e}")
        set_output("changed", "false")
        set_output("reason", f"anthropic-call-fehler: {type(e).__name__}")
        return 4
    log(f"      stop_reason: {stop_reason}")
    log(f"      usage: {usage}")
    log(f"      Rohantwort: {len(raw_text)} Zeichen.")
    new_data_js = strip_code_fence(raw_text)
    log(f"      Nach Code-Fence-Strip: {len(new_data_js)} Zeichen.")

    log("[6/7] Ausgabe validieren ...")
    ok, msg = validate_data_js(new_data_js)
    if not ok:
        log(f"      FEHLER: {msg}")
        log("      Antwort (erste 1200 Zeichen):")
        log(raw_text[:1200] if raw_text else "(leer)")
        set_output("changed", "false")
        set_output("reason", f"validierung-fehlgeschlagen: {msg}")
        return 3

    if new_data_js.strip() == current_data_js.strip():
        log("      data.js unveraendert.")
        set_output("changed", "false")
        set_output("reason", "keine-aenderung")
        return 0

    log("[7/7] Neue data.js schreiben ...")
    if not new_data_js.endswith("\n"):
        new_data_js += "\n"
    DATA_JS.write_text(new_data_js, encoding="utf-8")
    set_output("changed", "true")
    set_output("reason", "erfolgreich")
    log("Fertig.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
